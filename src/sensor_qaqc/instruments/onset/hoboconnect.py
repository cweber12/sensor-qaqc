"""Reading a HOBOconnect workbook export (#3).

Everything here was verified against the recovered original export
(``docs/data/yellow_buoy_temps.xlsx``): three sheets, ``Data`` with an integer
sample number, a naive Excel serial timestamp and the value, ``Details`` as
key/value rows under section headings with every value stored as a *string*,
and ``Events`` as one column per event type. The Google Sheets round-trip that
was mistaken for an export is kept as a corrupt fixture and taught nothing.

This module owns Onset's vocabulary and nothing else: ``Fixed - Normal``
becomes ``fixed``, ``0 hour 10 minutes 0 seconds`` becomes ``600``, ``°F``
becomes the UDUNITS-2 symbol ``degF``. Table shape comes from ``sources.yaml``
and generic row parsing from ``tables``, so a second workbook vendor reuses
both without inheriting these words.

Two facts the export states twice are checked against each other, because
agreement is free and disagreement means the parse is wrong: the unit appears
in the Data header and in the Details series declaration, and the zone label
appears in the Data header and on each published sample time. Neither check
overlaps the Details statistics gate, which compares the numbers.
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import TYPE_CHECKING

import numpy as np
import openpyxl

from sensor_qaqc.core.records import LoggingMode
from sensor_qaqc.instruments.extraction import (
    ExtractedMetadata,
    Extraction,
    PublishedStatistics,
)
from sensor_qaqc.instruments.tables import parse_data, parse_details
from sensor_qaqc.instruments.timezones import to_utc

if TYPE_CHECKING:
    from collections.abc import Sequence
    from pathlib import Path

    from sensor_qaqc.instruments.sources import SourceFormat
    from sensor_qaqc.instruments.tables import DetailsTable

# UDUNITS-2 symbols for the degree signs Onset writes in a column header.
UNIT_SYMBOLS = {"\N{DEGREE SIGN}F": "degF", "\N{DEGREE SIGN}C": "degC"}
# "Fixed - Normal": the mode, then the sampling style. Burst arrives with the
# first export that carries it, together with its LoggingMode member.
LOGGING_MODES = {"Fixed": LoggingMode.FIXED}
INTERVAL = re.compile(
    r"^(?P<hours>\d+)\s*hours?\s+(?P<minutes>\d+)\s*minutes?\s+(?P<seconds>\d+)\s*seconds?$"
)
# "2026/07/11 07:00:00 PDT" - the stamp, then the same label the Data header
# declares. Parsed with the label split off, never with a zone-aware format
# code: %Z accepts an abbreviation and then quietly ignores it.
DETAILS_TIME = "%Y/%m/%d %H:%M:%S"

DEVICE_INFO = "Device Info"
DEPLOYMENT_INFO = "Deployment Info"
SERIES_STATISTICS = "Series Statistics"
LOCATION_OFF = "Off"


class HOBOconnectReader:
    """Reads one HOBOconnect workbook into an ``Extraction``.

    The shape is injected rather than imported: the same reader runs against a
    corrected ``sources.yaml`` entry without a code change, which is what makes
    the catalogue the description of the format rather than documentation of it.
    """

    format_id = "hoboconnect_xlsx"

    def __init__(self, source_format: SourceFormat) -> None:
        if source_format.details is None:
            raise ValueError(f"{source_format.format_id} declares no details table shape")
        self._format = source_format
        self._details_spec = source_format.details

    def read(self, path: Path) -> Extraction:
        """Parse the workbook. Nothing is trimmed, masked or converted."""
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            # Read both sheets out before closing: a read_only worksheet is a
            # cursor into the open file, not a table already in memory.
            data = parse_data(self._rows(workbook, "data"), self._format.data)
            details = parse_details(self._rows(workbook, "details"), self._details_spec)
        finally:
            workbook.close()

        notes = list(data.notes)
        units = self._units(data.unit)
        self._check_declared_unit(details, units)
        published, published_notes = self._statistics(details, units, data.timezone_label)
        return Extraction(
            format_id=self.format_id,
            timestamps=to_utc(data.timestamps, data.timezone_label),
            values=np.asarray(data.values, dtype=np.float64),
            metadata=self._metadata(details, units, data.timezone_label, notes),
            published=published,
            notes=tuple(notes + published_notes),
        )

    def _rows(self, workbook: openpyxl.Workbook, table: str) -> list[Sequence[object]]:
        """Return the rows of the sheet ``sources.yaml`` puts this table in."""
        where = self._format.tables.get(table)
        if where is None:
            raise ValueError(f"{self.format_id} does not say which sheet holds the {table} table")
        if where not in workbook.sheetnames:
            raise ValueError(
                f"the workbook has no {where!r} sheet for the {table} table;"
                f" it has: {', '.join(workbook.sheetnames)}"
            )
        return list(workbook[where].iter_rows(values_only=True))

    def _units(self, raw: str) -> str:
        if raw not in UNIT_SYMBOLS:
            known = ", ".join(sorted(UNIT_SYMBOLS))
            raise ValueError(
                f"the data header declares the unit {raw!r}, which this reader does not"
                f" normalise; it knows: {known}"
            )
        return UNIT_SYMBOLS[raw]

    def _check_declared_unit(self, details: DetailsTable, units: str) -> None:
        if details.series is None:
            raise ValueError("the details table declares no series, so its unit cannot be checked")
        declared = self._format.data.value_column.match(details.series)
        if declared is None:
            raise ValueError(
                f"the details series declaration {details.series!r} does not have the shape"
                " sources.yaml gives a value column"
            )
        series_units = self._units(str(declared.group("unit")))
        if series_units != units:
            raise ValueError(
                f"the data header declares {units} but the details series declares"
                f" {series_units}; the export states the unit twice and the parse cannot"
                " honour both"
            )

    def _metadata(
        self, details: DetailsTable, units: str, label: str, notes: list[str]
    ) -> ExtractedMetadata:
        location = details.optional(DEPLOYMENT_INFO, "Location")
        if location is not None and location != LOCATION_OFF:
            # No export with Location switched on has been seen, so its format
            # is unverified: guessing at it would invent a position, and
            # ignoring it would lose one. TODO(verify) against a real export.
            notes.append(
                f"Location is {location!r}; the format of a set Location is unverified,"
                " so no position was extracted - supply one if it is needed"
            )
        return ExtractedMetadata(
            product=details.value(DEVICE_INFO, "Product"),
            serial=details.value(DEVICE_INFO, "Serial Number"),
            deployment_number=_int(
                details.value(DEPLOYMENT_INFO, "Deployment Number"), "Deployment Number"
            ),
            interval_s=_interval_seconds(details.value(DEPLOYMENT_INFO, "Logging Interval")),
            units=units,
            source_timezone_label=label,
            firmware=details.optional(DEVICE_INFO, "Firmware Version"),
            logging_mode=_logging_mode(details.value(DEPLOYMENT_INFO, "Logging Mode")),
        )

    def _statistics(
        self, details: DetailsTable, units: str, label: str
    ) -> tuple[PublishedStatistics | None, list[str]]:
        published = details.section(SERIES_STATISTICS)
        if not published:
            return None, [
                (
                    "the details table publishes no series statistics,"
                    " so the checksum gate has nothing to reproduce"
                )
            ]
        return (
            PublishedStatistics(
                samples=_int(published["Samples"], "Samples"),
                maximum=_float(published["Max"], "Max"),
                minimum=_float(published["Min"], "Min"),
                average=_float(published["Avg"], "Avg"),
                std_dev=_float(published["Std Dev"], "Std Dev"),
                first_sample_time=_sample_time(published["First Sample Time"], label),
                last_sample_time=_sample_time(published["Last Sample Time"], label),
                units=units,
            ),
            [],
        )


def _int(raw: str, name: str) -> int:
    try:
        return int(raw)
    except ValueError as error:
        raise ValueError(f"the details value for {name} is {raw!r}, not a whole number") from error


def _float(raw: str, name: str) -> float:
    try:
        return float(raw)
    except ValueError as error:
        raise ValueError(f"the details value for {name} is {raw!r}, not a number") from error


def _interval_seconds(raw: str) -> int:
    matched = INTERVAL.match(raw.strip())
    if matched is None:
        raise ValueError(
            f"the details value for Logging Interval is {raw!r}, which does not have the"
            " shape '<n> hour <n> minutes <n> seconds'"
        )
    hours, minutes, seconds = (int(matched.group(part)) for part in ("hours", "minutes", "seconds"))
    total = hours * 3600 + minutes * 60 + seconds
    if total <= 0:
        raise ValueError(f"the details value for Logging Interval is {raw!r}, which is no interval")
    return total


def _logging_mode(raw: str) -> LoggingMode:
    mode = raw.split("-", maxsplit=1)[0].strip()
    if mode not in LOGGING_MODES:
        known = ", ".join(sorted(LOGGING_MODES))
        raise ValueError(
            f"the details value for Logging Mode is {raw!r}; this reader normalises: {known}."
            " A mode arrives with the export that carries it, so its timestamps can be"
            " verified rather than assumed"
        )
    return LOGGING_MODES[mode]


def _sample_time(raw: str, label: str) -> object:
    stamp, _, declared = raw.rpartition(" ")
    if declared != label:
        raise ValueError(
            f"the published sample time {raw!r} declares the zone {declared!r} but the data"
            f" header declares {label!r}; the export states the zone twice and they disagree"
        )
    try:
        naive = datetime.strptime(stamp, DETAILS_TIME)  # noqa: DTZ007 - naive local by construction; the label is checked above and applied by to_utc
    except ValueError as error:
        raise ValueError(f"the published sample time {raw!r} is not a readable stamp") from error
    return to_utc([naive], label)[0]
