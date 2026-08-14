"""Getting a source's tables out of whatever holds them (#3).

A source is *tables*, not a file: a data table, an optional events table and an
optional details table, living in the sheets of one workbook or in sibling CSV
files. This module is the only place that knows which. Everything above it -
row parsing, vendor vocabulary, assembly - sees rows and nothing else, which is
what lets the same Onset reader read the same tables out of either container.

Cells arrive typed from a workbook and as text from a CSV, so each container
also supplies the coercion the row parsers use. A CSV's timestamp format is
*declared* in ``sources.yaml`` rather than inferred: guessing at a date format
is how a day becomes a month for the twelve days of a year where both readings
parse, and the tool would never know.

A declared table that is absent is refused, not shrugged at. The format says
the source has one; a bare CSV that genuinely has no metadata is a different
format entry, and its ingest reports the missing gate rather than hiding it.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from typing import TYPE_CHECKING, Protocol

import openpyxl

if TYPE_CHECKING:
    from collections.abc import Mapping, Sequence
    from pathlib import Path

    from sensor_qaqc.instruments.sources import SourceFormat

XLSX = "xlsx"
CSV = "csv"


class CellCoercion(Protocol):
    """How a container's cells become the types the row parsers expect."""

    def timestamp(self, raw: object) -> object:
        """Return the cell as a datetime, or raise ValueError."""
        ...

    def number(self, raw: object) -> object:
        """Return the cell as a number, or raise ValueError."""
        ...


@dataclass(frozen=True)
class TypedCells:
    """A workbook cell already carries its type; nothing to do."""

    def timestamp(self, raw: object) -> object:
        return raw

    def number(self, raw: object) -> object:
        return raw


@dataclass(frozen=True)
class TextCells:
    """Every CSV cell is text, so the stamp format has to be declared."""

    timestamp_format: str

    def timestamp(self, raw: object) -> object:
        return datetime.strptime(str(raw), self.timestamp_format)  # noqa: DTZ007 - naive local; assemble localises

    def number(self, raw: object) -> object:
        return float(str(raw))


# A module-level singleton, not a default-argument call: the row parsers
# take it as a default, and a fresh instance per call would be built for every
# table read.
TYPED_CELLS = TypedCells()


@dataclass(frozen=True)
class LoadedSource:
    """Every declared table's rows, and how to read this container's cells."""

    tables: Mapping[str, list[Sequence[object]]]
    cells: CellCoercion


def read_tables(source_format: SourceFormat, path: Path) -> LoadedSource:
    """Read every table the format declares out of the container at ``path``."""
    if source_format.container == XLSX:
        return LoadedSource(tables=_from_workbook(source_format, path), cells=TypedCells())
    if source_format.container == CSV:
        if source_format.timestamp_format is None:  # pragma: no cover - the catalogue refuses first
            raise ValueError(f"{source_format.format_id} declares no timestamp_format")
        return LoadedSource(
            tables=_from_csv_files(source_format, path),
            cells=TextCells(source_format.timestamp_format),
        )
    raise ValueError(
        f"{source_format.format_id} declares the container {source_format.container!r},"
        " which nothing knows how to open"
    )


def _from_workbook(source_format: SourceFormat, path: Path) -> dict[str, list[Sequence[object]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        missing = [
            f"{where!r} for the {table} table"
            for table, where in source_format.tables.items()
            if where not in workbook.sheetnames
        ]
        if missing:
            raise ValueError(
                f"the workbook has no sheet {' or '.join(missing)};"
                f" it has: {', '.join(workbook.sheetnames)}"
            )
        # Read each sheet out before closing: a read_only worksheet is a
        # cursor into the open file, not a table already in memory.
        return {
            table: list(workbook[where].iter_rows(values_only=True))
            for table, where in source_format.tables.items()
        }
    finally:
        workbook.close()


def _from_csv_files(source_format: SourceFormat, path: Path) -> dict[str, list[Sequence[object]]]:
    # Pointing at the data file finds its siblings; pointing at the directory
    # finds them all by their declared names.
    directory = path if path.is_dir() else path.parent
    located = {
        table: (path if table == "data" and not path.is_dir() else directory / name)
        for table, name in source_format.tables.items()
    }
    # Every absent table named at once: told about them one at a time, an
    # operator assembles a bundle one re-run per missing file.
    missing = [
        f"{where.name!r} for the {table} table"
        for table, where in located.items()
        if not where.is_file()
    ]
    if missing:
        raise ValueError(
            f"nothing at {directory.as_posix()} is {' or '.join(missing)};"
            f" {source_format.format_id} declares them"
        )
    return {table: _read_csv(where) for table, where in located.items()}


def _read_csv(path: Path) -> list[Sequence[object]]:
    # utf-8-sig, not utf-8: a byte-order mark is unambiguous when present, and
    # left in place it becomes part of the first header cell, which then
    # matches no column pattern for a reason nobody would guess.
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [[cell if cell != "" else None for cell in row] for row in csv.reader(handle)]
