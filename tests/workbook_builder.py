"""Build HOBOconnect-shaped workbooks for tests that the fixtures cannot cover.

The pristine export is the only real HOBOconnect file available (no app access,
#3 plan). It contains no ``Power Warn``, no ``Safe Shutdown``, no set
``Location`` and - being correct - no corruption. Every test of those needs a
synthetic workbook, so this builder exists to write one whose *shape* is copied
from the real file: the same sheet names, the same header strings, Details as
key/value rows under section headers, statistics as 2-decimal strings.

By default the Details statistics are computed from the data rows, so the
workbook is self-consistent and the checksum gate passes on it. That is what
makes a corrupted variant a one-argument change, and it is why the corruption
tests are honest: the only difference between the passing and failing case is
the corruption itself.
"""

from __future__ import annotations

import statistics
from typing import TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from collections.abc import Mapping, Sequence
    from datetime import datetime
    from pathlib import Path

UNIT = "\N{DEGREE SIGN}F"
SERIES = "Tidbit 1"
ZONE = "PDT"

# Section -> ordered keys, as the real Details sheet lays them out.
DEFAULT_DETAILS: dict[str, dict[str, str]] = {
    "App Info": {"App Name": "HOBOconnect", "Version": "2.11.0 (1783017847)"},
    "Device Info": {
        "Product": "MX2204",
        "Serial Number": "22506632",
        "Firmware Version": "62.140",
        "Manufacturer": "Onset Computer Corporation",
    },
    "Deployment Info": {
        "Name": f"{SERIES} (22506632)",
        "Location": "Off",
        "Deployment Number": "3",
        "Logging Interval": "0 hour 10 minutes 0 seconds",
        "Logging Mode": "Fixed - Normal",
        # The trailing space is Onset's own, present in the vendor file too.
        "Stop Logging": "Stop When Memory Fills ",
    },
    "Alarm Settings": {"High Alarm Value": "Not Configured", "Low Alarm Value": "Not Configured"},
}

STATISTICS_SECTION = "Series Statistics"
# Sections the vendor file places *after* the series declaration row.
PER_SERIES_SECTIONS = ("Alarm Settings", STATISTICS_SECTION)


def statistics_for(samples: Sequence[tuple[datetime, float]], zone: str = ZONE) -> dict[str, str]:
    """Compute the Details statistics Onset would publish for these samples."""
    values = [value for _, value in samples]
    return {
        "Samples": str(len(samples)),
        "Max": f"{max(values):.2f}",
        "Min": f"{min(values):.2f}",
        "Avg": f"{statistics.fmean(values):.2f}",
        "Std Dev": f"{statistics.pstdev(values):.2f}",
        "First Sample Time": _stamp(samples[0][0], zone),
        "Last Sample Time": _stamp(samples[-1][0], zone),
    }


def _stamp(when: datetime, zone: str = ZONE) -> str:
    return f"{when:%Y/%m/%d %H:%M:%S} {zone}"


def write_workbook(  # noqa: PLR0913 - each argument names one thing a test corrupts
    path: Path,
    samples: Sequence[tuple[datetime, float]],
    *,
    details: Mapping[str, Mapping[str, str]] | None = None,
    published: Mapping[str, str] | None = None,
    unit: str = UNIT,
    zone: str = ZONE,
    events: Sequence[tuple[int, datetime, str]] = (),
    event_columns: Sequence[str] = (),
    data_header: Sequence[str] | None = None,
    blank_rows_before_data: int = 0,
) -> Path:
    """Write a workbook shaped like a HOBOconnect export. Returns ``path``."""
    workbook = openpyxl.Workbook()
    data = workbook.active
    data.title = "Data"
    default_header = ["#", f"Date-Time ({zone})", f"{SERIES} , {unit}"]
    data.append(list(data_header) if data_header is not None else default_header)
    for _ in range(blank_rows_before_data):
        data.append([None, None, None])
    for number, (when, value) in enumerate(samples, start=1):
        data.append([number, when, value])

    sheet = workbook.create_sheet("Events")
    sheet.append(["#", f"Date-Time ({zone})", *event_columns])
    for number, when, column in events:
        row: list[object] = [number, when, *[None] * len(event_columns)]
        row[2 + list(event_columns).index(column)] = "Logged"
        sheet.append(row)

    sections: dict[str, Mapping[str, str]] = dict(
        details if details is not None else DEFAULT_DETAILS
    )
    if published is not None:
        sections[STATISTICS_SECTION] = published
    elif samples:
        sections[STATISTICS_SECTION] = statistics_for(samples, zone)
    _write_details(workbook.create_sheet("Details"), sections, unit=unit)

    workbook.save(path)
    return path


def _write_details(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    sections: Mapping[str, Mapping[str, str]],
    *,
    unit: str,
) -> None:
    sheet.append(["Details"])
    sheet.append(["Devices"])
    declared = False
    for section, entries in sections.items():
        if not declared and section in PER_SERIES_SECTIONS:
            # The series declaration separates the deployment sections from the
            # per-series ones, exactly as the vendor file lays it out.
            sheet.append([f"Series : {SERIES} , {unit}"])
            declared = True
        sheet.append([None, section])
        for key, value in entries.items():
            sheet.append([None, None, key, value])
