"""The HOBOconnect reader, against the pristine export and synthetic corruptions (#3).

Every claim about the vendor format here was read off
``docs/data/yellow_buoy_temps.xlsx`` (SHA-256 e5f6676e...), the recovered
original export - not off the Google Sheets round-trip that was mistaken for
one, and not off the prototype.

The reader's job ends at "what the file says". It does not trim, mask, convert
units or judge; the checksum gate and the checks do that.
"""

from __future__ import annotations

import math
from datetime import datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
import pandas as pd
import pytest

from sensor_qaqc.core.records import LoggingMode
from sensor_qaqc.instruments.extraction import SuppliedMetadata, assemble
from sensor_qaqc.instruments.onset.hoboconnect import HOBOconnectReader
from sensor_qaqc.instruments.sensors import load_sensor_catalogue
from sensor_qaqc.instruments.sources import load_source_catalogue
from sensor_qaqc.instruments.tables import parse_details
from workbook_builder import DEFAULT_DETAILS, write_workbook

if TYPE_CHECKING:
    from sensor_qaqc.instruments.extraction import Extraction

PRISTINE = Path(__file__).resolve().parents[1] / "docs" / "data" / "yellow_buoy_temps.xlsx"
EDITED = Path(__file__).resolve().parent / "data" / "yellow_buoy_temps_edited.xlsx"

# Published in Details, verified by reading the file: 3,029 samples on a
# complete 10-minute grid from 07:00 PDT on 2026-07-11 to 07:40 PDT on 08-01.
PRISTINE_SAMPLES = 3029
DEPLOYMENT_NUMBER = 3
INTERVAL_S = 600
FOUR_ROWS = 4
EDITED_SAMPLES = 3022
# The reader reports the source's own frame; assemble is what makes it UTC.
FIRST_LOCAL = datetime(2026, 7, 11, 7, 0)  # noqa: DTZ001 - naive local, as the sheet stores it
LAST_LOCAL = datetime(2026, 8, 1, 7, 40)  # noqa: DTZ001 - naive local, as the sheet stores it
FIRST_UTC = pd.Timestamp("2026-07-11T14:00:00Z")


def _reader() -> HOBOconnectReader:
    return HOBOconnectReader(load_source_catalogue().for_format("hoboconnect_xlsx"))


def _read(path: Path) -> Extraction:
    return _reader().read(path)


def _samples(n: int, *, start: datetime | None = None) -> list[tuple[datetime, float]]:
    first = start if start is not None else datetime(2026, 7, 11, 7, 0)  # noqa: DTZ001 - naive local, as the sheet stores it
    return [(first + timedelta(minutes=10 * k), 70.0 + k / 100) for k in range(n)]


# --- The pristine export, read as the vendor wrote it. ---


def test_the_reader_finds_every_sample_the_details_sheet_claims() -> None:
    extraction = _read(PRISTINE)
    assert len(extraction.timestamps) == PRISTINE_SAMPLES
    assert extraction.timestamps[0] == FIRST_LOCAL
    assert extraction.timestamps[-1] == LAST_LOCAL


def test_the_unit_and_the_zone_label_are_read_off_the_header() -> None:
    # Both live only in the column headers ("Tidbit 1 , °F", "Date-Time (PDT)").
    # Assuming either is how a record ends up scaled or shifted with nothing
    # to notice it.
    metadata = _read(PRISTINE).metadata
    assert metadata.units == "degF"
    assert metadata.source_timezone_label == "PDT"


def test_the_identity_and_logging_metadata_come_from_details() -> None:
    metadata = _read(PRISTINE).metadata
    assert metadata.product == "MX2204"
    assert metadata.serial == "22506632"
    assert metadata.firmware == "62.140"
    assert metadata.deployment_number == DEPLOYMENT_NUMBER
    assert metadata.interval_s == INTERVAL_S
    assert metadata.logging_mode is LoggingMode.FIXED


def test_a_location_switched_off_yields_no_position_rather_than_a_guess() -> None:
    metadata = _read(PRISTINE).metadata
    assert metadata.latitude is None
    assert metadata.longitude is None


def test_the_published_statistics_are_carried_verbatim() -> None:
    published = _read(PRISTINE).published
    assert published is not None
    assert published.samples == PRISTINE_SAMPLES
    assert published.maximum == pytest.approx(75.35)
    assert published.minimum == pytest.approx(58.60)
    assert published.average == pytest.approx(70.84)
    assert published.std_dev == pytest.approx(2.38)
    assert published.first_sample_time == FIRST_LOCAL
    assert published.last_sample_time == LAST_LOCAL
    assert published.units == "degF"


def test_details_values_are_stripped_of_onsets_trailing_space() -> None:
    # 'Stop When Memory Fills ' carries its trailing space in the vendor file
    # itself, so this is the real quirk and not an artefact of any round-trip.
    workbook = openpyxl.load_workbook(PRISTINE, read_only=True, data_only=True)
    spec = load_source_catalogue().for_format("hoboconnect_xlsx").details
    assert spec is not None
    details = parse_details(workbook["Details"].iter_rows(values_only=True), spec)
    workbook.close()
    assert details.value("Deployment Info", "Stop Logging") == "Stop When Memory Fills"
    assert details.series == "Tidbit 1 , \N{DEGREE SIGN}F"


def test_the_pristine_export_assembles_into_a_complete_canonical_record() -> None:
    record = assemble(
        _read(PRISTINE),
        SuppliedMetadata(variable="sea_water_temperature"),
        sensors=load_sensor_catalogue(),
    )
    assert record.deployment_id == "22506632-3"
    assert record.units == "degF"
    assert record.n_valid == PRISTINE_SAMPLES
    assert record.gap_fraction == 0.0
    assert record.duration == timedelta(days=21, minutes=40)
    assert record.series.index[0] == FIRST_UTC


# --- The Sheets-edited copy: it parses, and the gate is what refuses it (#3 slice 4). ---


def test_the_edited_copy_parses_its_remaining_rows_and_reports_the_blanks() -> None:
    extraction = _read(EDITED)
    assert len(extraction.timestamps) == EDITED_SAMPLES
    assert any("blank" in note for note in extraction.notes)


# --- Anything the reader cannot place refuses; nothing is dropped quietly. ---


def test_a_sample_number_that_jumps_refuses(tmp_path: Path) -> None:
    # A row removed from the middle leaves the numbering with a hole. The
    # statistics gate would catch it too, but only where a format publishes
    # them; this is the check that does not depend on that.
    path = write_workbook(tmp_path / "gap.xlsx", _samples(4))
    workbook = openpyxl.load_workbook(path)
    workbook["Data"].delete_rows(4)
    workbook.save(path)
    with pytest.raises(ValueError, match=r"sample number.*4.*3"):
        _read(path)


def test_a_row_with_a_value_but_no_timestamp_refuses(tmp_path: Path) -> None:
    path = write_workbook(tmp_path / "no_stamp.xlsx", _samples(4))
    workbook = openpyxl.load_workbook(path)
    workbook["Data"].cell(row=3, column=2).value = None
    workbook.save(path)
    with pytest.raises(ValueError, match="no timestamp"):
        _read(path)


def test_a_row_with_a_timestamp_but_no_value_is_a_gap_not_an_error(tmp_path: Path) -> None:
    # A logged sample with no reading is missing data, which the masking
    # contract represents as NaN in place - not as a row to drop.
    path = write_workbook(tmp_path / "no_value.xlsx", _samples(4))
    workbook = openpyxl.load_workbook(path)
    workbook["Data"].cell(row=3, column=3).value = None
    workbook.save(path)
    extraction = _read(path)
    readings = extraction.values  # noqa: PD011 - an Extraction is not a DataFrame
    assert len(readings) == FOUR_ROWS
    assert math.isnan(readings[1])


def test_a_zone_label_nothing_can_resolve_refuses_at_assembly(tmp_path: Path) -> None:
    # "PDT" is an abbreviation, not a zone. The reader reports whatever label
    # the header declared - that is what the file says - and assembly is where
    # a label must resolve to an offset or refuse. Nothing assumes one.
    path = write_workbook(tmp_path / "zone.xlsx", _samples(3), zone="XYZ")
    extraction = _read(path)
    assert extraction.metadata.source_timezone_label == "XYZ"
    with pytest.raises(LookupError, match=r"'XYZ'.*PDT"):
        assemble(
            extraction,
            SuppliedMetadata(variable="sea_water_temperature"),
            sensors=load_sensor_catalogue(),
        )


def test_a_logging_mode_the_vocabulary_does_not_have_refuses(tmp_path: Path) -> None:
    details = _details_with({"Logging Mode": "Burst - Normal"})
    path = write_workbook(tmp_path / "burst.xlsx", _samples(3), details=details)
    with pytest.raises(ValueError, match=r"Logging Mode.*Burst"):
        _read(path)


def test_a_logging_interval_that_does_not_parse_refuses(tmp_path: Path) -> None:
    details = _details_with({"Logging Interval": "every now and then"})
    path = write_workbook(tmp_path / "interval.xlsx", _samples(3), details=details)
    with pytest.raises(ValueError, match="Logging Interval"):
        _read(path)


def test_a_unit_the_reader_cannot_normalise_refuses(tmp_path: Path) -> None:
    path = write_workbook(tmp_path / "unit.xlsx", _samples(3), unit="Kelvins")
    with pytest.raises(ValueError, match="Kelvins"):
        _read(path)


def test_details_and_data_disagreeing_about_the_unit_refuses(tmp_path: Path) -> None:
    # The unit appears twice in the export - the Data header and the Series
    # declaration. Two statements of one fact is a free checksum on unit
    # handling, and disagreement means the parse is wrong somewhere.
    path = write_workbook(tmp_path / "units.xlsx", _samples(3))
    workbook = openpyxl.load_workbook(path)
    data = workbook["Data"]
    data.cell(row=1, column=3).value = "Tidbit 1 , \N{DEGREE SIGN}C"
    workbook.save(path)
    with pytest.raises(ValueError, match=r"degC.*degF|degF.*degC"):
        _read(path)


def test_a_missing_sheet_refuses_naming_what_the_workbook_has(tmp_path: Path) -> None:
    path = write_workbook(tmp_path / "no_details.xlsx", _samples(3))
    workbook = openpyxl.load_workbook(path)
    del workbook["Details"]
    workbook.save(path)
    with pytest.raises(ValueError, match=r"Details.*Data"):
        _read(path)


def test_a_second_series_column_refuses_rather_than_taking_the_first(tmp_path: Path) -> None:
    path = write_workbook(
        tmp_path / "two_series.xlsx",
        _samples(3),
        data_header=[
            "#",
            "Date-Time (PDT)",
            "Tidbit 1 , \N{DEGREE SIGN}F",
            "Tidbit 2 , \N{DEGREE SIGN}F",
        ],
    )
    with pytest.raises(ValueError, match="Tidbit 2"):
        _read(path)


def test_a_location_that_is_set_is_reported_rather_than_parsed(tmp_path: Path) -> None:
    # No export with Location switched on has ever been seen, so its format is
    # unverified. Guessing at it would invent a position; ignoring it silently
    # would lose one. The reader says what it saw and supplies no position.
    details = _details_with({"Location": "32.8663, -117.2544"})
    path = write_workbook(tmp_path / "located.xlsx", _samples(3), details=details)
    extraction = _read(path)
    assert extraction.metadata.latitude is None
    assert any("Location" in note for note in extraction.notes)


def _details_with(overrides: dict[str, str]) -> dict[str, dict[str, str]]:
    details = {section: dict(entries) for section, entries in DEFAULT_DETAILS.items()}
    details["Deployment Info"].update(overrides)
    return details
