"""The event log is the audit trail, and its columns are dynamic (#3).

The export writes **one column per event type**, and only for types that
occurred: this deployment's log has ``Host Connected``, ``End of File`` and
``Started`` and nothing else. A parser addressing fixed positions would read
the wrong column on the next export, so the column set is discovered from the
header every time.

The pristine log holds five events end to end - ``Started`` at the first
sample, ``End of File`` at the last - so ``Power Warn``, ``Safe Shutdown``,
``Water Detect`` and ``New Interval`` can only be exercised with synthetic
sheets. That is what most of this file is.

Ingest surfaces events. What one means for a verdict is #7's decision, with an
injected, provenance-carrying policy - there is deliberately no judgement here.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
import pandas as pd
import pytest

from sensor_qaqc.core.records import EventType
from sensor_qaqc.instruments.extraction import SuppliedMetadata, assemble
from sensor_qaqc.instruments.onset.hoboconnect import HOBOconnectReader
from sensor_qaqc.instruments.sensors import load_sensor_catalogue
from sensor_qaqc.instruments.sources import load_source_catalogue
from workbook_builder import write_workbook

if TYPE_CHECKING:
    from sensor_qaqc.instruments.extraction import Extraction

PRISTINE = Path(__file__).resolve().parents[1] / "docs" / "data" / "yellow_buoy_temps.xlsx"
EDITED = Path(__file__).resolve().parent / "data" / "yellow_buoy_temps_edited.xlsx"
PRISTINE_EVENTS = 5


def _read(path: Path) -> Extraction:
    reader = HOBOconnectReader(load_source_catalogue().for_format("hoboconnect_xlsx"))
    return reader.read(path)


def _events(path: Path) -> tuple[tuple[str, str], ...]:
    extraction = _read(path)
    return tuple((event.event_type.value, str(event.at)) for event in extraction.events)


def _samples(n: int) -> list[tuple[datetime, float]]:
    first = datetime(2026, 7, 11, 7, 0)  # noqa: DTZ001 - naive local, as the sheet stores it
    return [(first + timedelta(minutes=10 * k), 70.0 + k / 100) for k in range(n)]


# --- The log the real export carries. ---


def test_the_pristine_log_is_read_end_to_end_in_file_order() -> None:
    assert [name for name, _ in _events(PRISTINE)] == [
        "host_connected",
        "started",
        "host_connected",
        "host_connected",
        "end_of_file",
    ]


def test_an_event_from_before_the_first_sample_is_kept() -> None:
    # The launch connection is 2026-07-10 15:26 PDT, the day before logging
    # started. Constraining events to the sample span would discard it.
    first = _events(PRISTINE)[0]
    assert first == ("host_connected", "2026-07-10 22:26:47+00:00")


def test_the_log_reaches_the_last_sample() -> None:
    # End of File at the last sample time is what makes this log complete
    # rather than truncated - the "exactly 1,000 rows" scare was the Google
    # Sheets grid in the edited copy, not an export limit.
    name, when = _events(PRISTINE)[-1]
    assert (name, when) == ("end_of_file", "2026-08-01 14:40:00+00:00")


def test_the_assembled_record_carries_the_log() -> None:
    record = assemble(
        _read(PRISTINE),
        SuppliedMetadata(variable="sea_water_temperature"),
        sensors=load_sensor_catalogue(),
    )
    assert len(record.events) == PRISTINE_EVENTS
    assert record.events[1].event_type is EventType.STARTED
    assert record.events[1].at == pd.Timestamp("2026-07-11T14:00:00Z")


def test_the_edited_copys_grid_padding_is_skipped_and_counted() -> None:
    # 994 styled-empty rows of Google Sheets grid. Skipping them silently is
    # how "the export truncates at 1,000 events" became a plausible story.
    extraction = _read(EDITED)
    assert len(extraction.events) == PRISTINE_EVENTS
    assert any("blank" in note and "event" in note for note in extraction.notes)


# --- Types the fixture cannot show, on synthetic sheets. ---


@pytest.mark.parametrize(
    ("column", "expected"),
    [
        ("Power Warn", EventType.POWER_WARN),
        ("Safe Shutdown", EventType.SAFE_SHUTDOWN),
        ("Water Detect", EventType.WATER_DETECT),
        ("New Interval", EventType.NEW_INTERVAL),
    ],
)
def test_the_types_this_deployment_never_logged_normalise(
    tmp_path: Path, column: str, expected: EventType
) -> None:
    when = datetime(2026, 7, 11, 7, 5)  # noqa: DTZ001 - naive local, as the sheet stores it
    path = write_workbook(
        tmp_path / "typed.xlsx",
        _samples(3),
        events=[(1, when, column)],
        event_columns=[column],
    )
    extraction = _read(path)
    assert [event.event_type for event in extraction.events] == [expected]


def test_the_columns_are_found_by_name_not_by_position(tmp_path: Path) -> None:
    # Same two events, opposite column order. A parser reading positions would
    # swap them and nothing downstream would notice.
    when = datetime(2026, 7, 11, 7, 5)  # noqa: DTZ001 - naive local, as the sheet stores it
    forward = write_workbook(
        tmp_path / "forward.xlsx",
        _samples(3),
        events=[(1, when, "Started"), (2, when, "Power Warn")],
        event_columns=["Started", "Power Warn"],
    )
    reversed_ = write_workbook(
        tmp_path / "reversed.xlsx",
        _samples(3),
        events=[(1, when, "Started"), (2, when, "Power Warn")],
        event_columns=["Power Warn", "Started"],
    )
    assert _events(forward) == _events(reversed_)


def test_an_event_type_the_vocabulary_does_not_have_refuses(tmp_path: Path) -> None:
    # Refusing rather than dropping: an unrecognised entry in an audit trail is
    # the last thing to pass over quietly, and growing the vocabulary is a
    # one-line diff in core/records.py.
    when = datetime(2026, 7, 11, 7, 5)  # noqa: DTZ001 - naive local, as the sheet stores it
    path = write_workbook(
        tmp_path / "unknown.xlsx",
        _samples(3),
        events=[(1, when, "Bluetooth Reset")],
        event_columns=["Bluetooth Reset"],
    )
    with pytest.raises(ValueError, match=r"Bluetooth Reset"):
        _read(path)


def test_a_marked_row_with_no_timestamp_refuses(tmp_path: Path) -> None:
    when = datetime(2026, 7, 11, 7, 5)  # noqa: DTZ001 - naive local, as the sheet stores it
    path = write_workbook(
        tmp_path / "no_stamp.xlsx",
        _samples(3),
        events=[(1, when, "Started")],
        event_columns=["Started"],
    )
    workbook = openpyxl.load_workbook(path)
    workbook["Events"].cell(row=2, column=2).value = None
    workbook.save(path)
    with pytest.raises(ValueError, match="no timestamp"):
        _read(path)


def test_a_row_that_marks_nothing_refuses(tmp_path: Path) -> None:
    when = datetime(2026, 7, 11, 7, 5)  # noqa: DTZ001 - naive local, as the sheet stores it
    path = write_workbook(
        tmp_path / "unmarked.xlsx",
        _samples(3),
        events=[(1, when, "Started")],
        event_columns=["Started"],
    )
    workbook = openpyxl.load_workbook(path)
    workbook["Events"].cell(row=2, column=3).value = None
    workbook.save(path)
    with pytest.raises(ValueError, match="marks no event"):
        _read(path)


def test_a_cell_holding_something_other_than_the_marker_refuses(tmp_path: Path) -> None:
    when = datetime(2026, 7, 11, 7, 5)  # noqa: DTZ001 - naive local, as the sheet stores it
    path = write_workbook(
        tmp_path / "odd_marker.xlsx",
        _samples(3),
        events=[(1, when, "Started")],
        event_columns=["Started"],
    )
    workbook = openpyxl.load_workbook(path)
    workbook["Events"].cell(row=2, column=3).value = "Maybe"
    workbook.save(path)
    with pytest.raises(ValueError, match=r"'Maybe'.*'Logged'"):
        _read(path)


def test_the_events_sheet_declaring_a_different_zone_refuses(tmp_path: Path) -> None:
    # A third statement of the zone, and the same free checksum: two sheets
    # written in different frames would shift the log against the samples.
    when = datetime(2026, 7, 11, 7, 5)  # noqa: DTZ001 - naive local, as the sheet stores it
    path = write_workbook(
        tmp_path / "zones.xlsx",
        _samples(3),
        events=[(1, when, "Started")],
        event_columns=["Started"],
    )
    workbook = openpyxl.load_workbook(path)
    workbook["Events"].cell(row=1, column=2).value = "Date-Time (PST)"
    workbook.save(path)
    with pytest.raises(ValueError, match=r"PST.*PDT|PDT.*PST"):
        _read(path)
